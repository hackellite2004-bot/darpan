"""
Report Generator for Darpan
Generates Excel reports with styled tables and comprehensive analysis.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Optional
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class ReportGenerator:
    """Generate comprehensive Excel reports for session analysis."""

    # Styling constants
    HEADER_FILL = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    HEADER_FONT = Font(bold=True, color='FFFFFF', size=12)
    SUBHEADER_FILL = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    SUBHEADER_FONT = Font(bold=True, color='000000', size=11)
    GOOD_FILL = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')
    WARNING_FILL = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
    BAD_FILL = PatternFill(start_color='F4B084', end_color='F4B084', fill_type='solid')
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LEFT_ALIGN = Alignment(horizontal='left', vertical='top', wrap_text=True)

    @staticmethod
    def generate_session_report(
        session_data: dict[str, Any],
        user_data: Optional[dict[str, Any]] = None,
        output_path: Optional[Path] = None,
    ) -> Path | None:
        """
        Generate a comprehensive Excel report for a session.
        
        Args:
            session_data: Session metrics and analysis
            user_data: User information
            output_path: Path to save the report
            
        Returns:
            Path to saved report or None if generation failed
        """
        if not HAS_OPENPYXL:
            print("openpyxl not installed. Install with: pip install openpyxl")
            return None

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Session Analysis"

            # Set column widths
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15

            row = 1

            # Title section
            ws.merge_cells(f'A{row}:D{row}')
            title_cell = ws[f'A{row}']
            title_cell.value = '📊 SESSION ANALYSIS REPORT'
            title_cell.font = Font(bold=True, size=16, color='FFFFFF')
            title_cell.fill = ReportGenerator.HEADER_FILL
            title_cell.alignment = ReportGenerator.CENTER_ALIGN
            ws.row_dimensions[row].height = 25
            row += 1

            # Metadata section
            ws[f'A{row}'].value = 'Date & Time:'
            ws[f'B{row}'].value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            row += 1

            if user_data:
                ws[f'A{row}'].value = 'User:'
                ws[f'B{row}'].value = user_data.get('name', 'Unknown')
                row += 1

                ws[f'A{row}'].value = 'Age Group:'
                ws[f'B{row}'].value = user_data.get('age_group', 'Unknown')
                row += 1

            row += 1

            # Overall Score section
            ws.merge_cells(f'A{row}:D{row}')
            section_cell = ws[f'A{row}']
            section_cell.value = 'OVERALL PERFORMANCE'
            section_cell.font = ReportGenerator.SUBHEADER_FONT
            section_cell.fill = ReportGenerator.SUBHEADER_FILL
            row += 1

            score = session_data.get('overall_score', session_data.get('confidence_score', 0))
            model_analysis = session_data.get('model_analysis') if isinstance(session_data.get('model_analysis'), dict) else {}
            grade = session_data.get('grade') or model_analysis.get('grade') or '-'
            confidence_raw = session_data.get('confidence_label', 'uncertain')
            if isinstance(confidence_raw, dict):
                confidence = str(confidence_raw.get('label', 'uncertain'))
            else:
                confidence = str(confidence_raw or 'uncertain')

            ws[f'A{row}'].value = 'Overall Score:'
            ws[f'B{row}'].value = f"{score}/100"
            ws[f'B{row}'].font = Font(bold=True, size=14)
            row += 1

            ws[f'A{row}'].value = 'Grade:'
            ws[f'B{row}'].value = grade
            ws[f'B{row}'].font = Font(bold=True, size=14, color='FFFFFF')
            grade_fill = ReportGenerator.GOOD_FILL if grade in ['A', 'B'] else ReportGenerator.WARNING_FILL if grade == 'C' else ReportGenerator.BAD_FILL
            ws[f'B{row}'].fill = grade_fill
            row += 1

            ws[f'A{row}'].value = 'Confidence Level:'
            ws[f'B{row}'].value = confidence.title()
            row += 1

            row += 1

            # Emotion Analysis section
            ws.merge_cells(f'A{row}:D{row}')
            section_cell = ws[f'A{row}']
            section_cell.value = 'EMOTION ANALYSIS'
            section_cell.font = ReportGenerator.SUBHEADER_FONT
            section_cell.fill = ReportGenerator.SUBHEADER_FILL
            row += 1

            emotions = {
                'Happy': session_data.get('emotion_happy', 0),
                'Neutral': session_data.get('emotion_neutral', 0),
                'Sad': session_data.get('emotion_sad', 0),
                'Anxious': session_data.get('emotion_anxious', 0),
                'Surprised': session_data.get('emotion_surprised', 0),
            }

            ws[f'A{row}'].value = 'Emotion'
            ws[f'B{row}'].value = 'Percentage'
            ws[f'C{row}'].value = 'Status'
            for cell in [ws[f'A{row}'], ws[f'B{row}'], ws[f'C{row}']]:
                cell.font = ReportGenerator.HEADER_FONT
                cell.fill = ReportGenerator.HEADER_FILL
                cell.alignment = ReportGenerator.CENTER_ALIGN
            row += 1

            for emotion, value in emotions.items():
                ws[f'A{row}'].value = emotion
                ws[f'B{row}'].value = f"{value:.1f}%"
                status = '✓' if value > 20 else '○'
                ws[f'C{row}'].value = status
                ws[f'C{row}'].alignment = ReportGenerator.CENTER_ALIGN
                row += 1

            row += 1

            # Speech Analysis section
            ws.merge_cells(f'A{row}:D{row}')
            section_cell = ws[f'A{row}']
            section_cell.value = 'SPEECH ANALYSIS'
            section_cell.font = ReportGenerator.SUBHEADER_FONT
            section_cell.fill = ReportGenerator.SUBHEADER_FILL
            row += 1

            wpm = session_data.get('wpm', 0)
            ws[f'A{row}'].value = 'Words Per Minute (WPM):'
            ws[f'B{row}'].value = f"{wpm:.0f}"
            wpm_fill = ReportGenerator.GOOD_FILL if 140 <= wpm <= 160 else ReportGenerator.WARNING_FILL if 120 <= wpm < 180 else ReportGenerator.BAD_FILL
            ws[f'B{row}'].fill = wpm_fill
            row += 1

            ws[f'A{row}'].value = 'WPM Status:'
            if wpm < 120:
                status = '🐢 Too Slow'
            elif wpm < 140:
                status = '↗ Below Optimal'
            elif wpm <= 160:
                status = '✓ Optimal'
            elif wpm < 180:
                status = '↖ Above Optimal'
            else:
                status = '🚀 Too Fast'
            ws[f'B{row}'].value = status
            row += 1

            fillers = session_data.get('filler_count', 0)
            ws[f'A{row}'].value = 'Filler Words Count:'
            ws[f'B{row}'].value = fillers
            ws[f'B{row}'].fill = ReportGenerator.GOOD_FILL if fillers < 5 else ReportGenerator.WARNING_FILL if fillers < 10 else ReportGenerator.BAD_FILL
            row += 1

            pauses = session_data.get('pause_count', 0)
            ws[f'A{row}'].value = 'Long Pauses (>2s):'
            ws[f'B{row}'].value = pauses
            row += 1

            row += 1

            # Eye Contact section
            ws.merge_cells(f'A{row}:D{row}')
            section_cell = ws[f'A{row}']
            section_cell.value = 'EYE CONTACT & GAZE'
            section_cell.font = ReportGenerator.SUBHEADER_FONT
            section_cell.fill = ReportGenerator.SUBHEADER_FILL
            row += 1

            eye_center = session_data.get('eye_center_pct', 0)
            ws[f'A{row}'].value = 'Eye Contact (Center):'
            ws[f'B{row}'].value = f"{eye_center:.1f}%"
            ws[f'B{row}'].fill = ReportGenerator.GOOD_FILL if eye_center > 50 else ReportGenerator.WARNING_FILL if eye_center > 30 else ReportGenerator.BAD_FILL
            row += 1

            blinks = session_data.get('blink_count', 0)
            ws[f'A{row}'].value = 'Blink Count:'
            ws[f'B{row}'].value = blinks
            row += 1

            row += 1

            # Posture & Gestures section
            ws.merge_cells(f'A{row}:D{row}')
            section_cell = ws[f'A{row}']
            section_cell.value = 'POSTURE & GESTURES'
            section_cell.font = ReportGenerator.SUBHEADER_FONT
            section_cell.fill = ReportGenerator.SUBHEADER_FILL
            row += 1

            posture = session_data.get('posture_score', 0)
            ws[f'A{row}'].value = 'Posture Score:'
            ws[f'B{row}'].value = f"{posture:.1f}%"
            ws[f'B{row}'].fill = ReportGenerator.GOOD_FILL if posture > 60 else ReportGenerator.WARNING_FILL if posture > 40 else ReportGenerator.BAD_FILL
            row += 1

            slouch = session_data.get('slouch_pct', 0)
            ws[f'A{row}'].value = 'Slouching:'
            ws[f'B{row}'].value = f"{slouch:.1f}%"
            row += 1

            gesture_pos = session_data.get('gesture_positive_pct', 0)
            ws[f'A{row}'].value = 'Positive Gestures:'
            ws[f'B{row}'].value = f"{gesture_pos:.1f}%"
            row += 1

            row += 1

            # Voice Characteristics section
            ws.merge_cells(f'A{row}:D{row}')
            section_cell = ws[f'A{row}']
            section_cell.value = 'VOICE CHARACTERISTICS'
            section_cell.font = ReportGenerator.SUBHEADER_FONT
            section_cell.fill = ReportGenerator.SUBHEADER_FILL
            row += 1

            voice_score = session_data.get('voice_score', 0)
            ws[f'A{row}'].value = 'Voice Score:'
            ws[f'B{row}'].value = f"{voice_score:.1f}"
            row += 1

            pitch_var = session_data.get('pitch_variation', 0)
            ws[f'A{row}'].value = 'Pitch Variation:'
            ws[f'B{row}'].value = f"{pitch_var:.2f}"
            ws[f'B{row}'].fill = ReportGenerator.GOOD_FILL if pitch_var > 0.5 else ReportGenerator.WARNING_FILL
            row += 1

            ws[f'A{row}'].value = 'Monotone:'
            ws[f'B{row}'].value = 'Yes' if session_data.get('is_monotone') else 'No'
            row += 1

            ws[f'A{row}'].value = 'Voice Energy:'
            ws[f'B{row}'].value = f"{session_data.get('voice_energy', 0):.2f}"
            row += 1

            row += 1

            # Session Duration
            ws.merge_cells(f'A{row}:D{row}')
            section_cell = ws[f'A{row}']
            section_cell.value = 'SESSION INFORMATION'
            section_cell.font = ReportGenerator.SUBHEADER_FONT
            section_cell.fill = ReportGenerator.SUBHEADER_FILL
            row += 1

            duration = session_data.get('duration_seconds', 0)
            ws[f'A{row}'].value = 'Duration:'
            ws[f'B{row}'].value = f"{duration:.1f} seconds ({duration/60:.1f} mins)"
            row += 1

            if output_path:
                wb.save(output_path)
                return output_path
            else:
                return None

        except Exception as e:
            print(f"Error generating report: {e}")
            return None

    @staticmethod
    def generate_multiple_session_report(
        sessions: list[dict[str, Any]],
        output_path: Optional[Path] = None,
    ) -> Path | None:
        """
        Generate a comparison report across multiple sessions.
        
        Args:
            sessions: List of session data dicts
            output_path: Path to save the report
            
        Returns:
            Path to saved report or None
        """
        if not HAS_OPENPYXL or len(sessions) < 2:
            return None

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Progress Overview"

            # Title
            ws.merge_cells('A1:D1')
            title_cell = ws['A1']
            title_cell.value = '📊 PROGRESS TRACKING REPORT'
            title_cell.font = Font(bold=True, size=16, color='FFFFFF')
            title_cell.fill = ReportGenerator.HEADER_FILL
            ws.row_dimensions[1].height = 25

            # Headers
            headers = ['Session', 'Date', 'Score', 'Grade', 'WPM', 'Eye Contact %', 'Posture %']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col)
                cell.value = header
                cell.font = ReportGenerator.HEADER_FONT
                cell.fill = ReportGenerator.HEADER_FILL
                cell.alignment = ReportGenerator.CENTER_ALIGN

            # Data rows
            for row_idx, session in enumerate(sessions, 4):
                ws.cell(row=row_idx, column=1).value = f"Session {row_idx - 3}"
                ws.cell(row=row_idx, column=2).value = session.get('date', 'N/A')
                
                score = session.get('overall_score', 0)
                ws.cell(row=row_idx, column=3).value = f"{score:.1f}"
                ws.cell(row=row_idx, column=3).fill = ReportGenerator.GOOD_FILL if score > 70 else ReportGenerator.WARNING_FILL if score > 50 else ReportGenerator.BAD_FILL
                
                ws.cell(row=row_idx, column=4).value = session.get('grade', '-')
                ws.cell(row=row_idx, column=5).value = f"{session.get('wpm', 0):.0f}"
                ws.cell(row=row_idx, column=6).value = f"{session.get('eye_center_pct', 0):.1f}%"
                ws.cell(row=row_idx, column=7).value = f"{session.get('posture_score', 0):.1f}%"

            # Set column widths
            for i in range(1, 8):
                ws.column_dimensions[get_column_letter(i)].width = 15

            if output_path:
                wb.save(output_path)
                return output_path

            return None

        except Exception as e:
            print(f"Error generating progress report: {e}")
            return None
